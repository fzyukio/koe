"""
Import syllables (not elements) from luscinia (after songs have been imported)
"""
import argparse
import re

import openpyxl
import psycopg2.extras
from progress.bar import Bar

name_regex = re.compile('(\w{3})_(\d{4})_(\d{2})_(\d{2})_([\w\d]+)_(\d+)_(\w+)\.(B|EX|VG|G|OK)(\.[^ ]*)?\.wav')
name_regex_ignore_case = \
    re.compile('(\w{3})_(\d{4})_(\d{2})_(\d{2})_([\w\d]+)_(\d+)_(\w+)\.(\w{1,2})(\.[^ ]*)?\.(wav|WAV)')

COLUMN_NAMES = ['Song name', 'Problem', 'Recommend name']
parser = argparse.ArgumentParser(description='Process some integers.')
parser.add_argument(
    '--db',
    action='store',
    dest='db',
    required=True,
    type=str,
    help='Database name',
)

parser.add_argument(
    '--port',
    action='store',
    dest='port',
    required=True,
    type=str,
    help='Port',
)

parser.add_argument(
    '--host',
    action='store',
    dest='host',
    required=True,
    type=str,
    help='Host',
)

args = parser.parse_args()
port = args.port
db = args.db
host = args.host

conn = None
wb = openpyxl.Workbook()
ws_problems = wb.create_sheet('Non-conforming Labels', 0)
ws_noproblems = wb.create_sheet('Conforming Labels', 1)

ws_problems.append(COLUMN_NAMES)
ws_noproblems.append(COLUMN_NAMES)

try:
    port = int(port)
    conn = psycopg2.connect(
        "dbname={} user=sa password='sa' host={} port={}".format(db, host, port))
    conn.set_client_encoding('LATIN1')

    cur = conn.cursor()

    cur.execute('select name from songdata')
    songs = cur.fetchall()
    nsongs = len(songs)

    bar = Bar('Checking file...', max=nsongs)
    inconsistency_count = 0
    for idx, song in enumerate(songs):
        song_name = song[0]

        song_name_cleaned = song_name.replace(' ', '').replace('$', '')

        song_name_is_correct = 'Yes'
        song_name_is_clean = song_name == song_name_cleaned

        hint = ''
        if not song_name_is_clean:
            hint += 'song_name contains $ or spaces, '

        recommend = ''
        if name_regex.match(song_name_cleaned):
            recommend = song_name_cleaned

        matches = name_regex.match(song_name)
        if matches is None:
            song_name_is_correct = 'No'

        matches_ignore_case = name_regex_ignore_case.match(song_name_cleaned)
        if matches_ignore_case:
            location = matches_ignore_case.group(1)
            year = matches_ignore_case.group(2)
            month = matches_ignore_case.group(3)
            date = matches_ignore_case.group(4)
            track_id = matches_ignore_case.group(5)
            track_number = matches_ignore_case.group(6)
            gender = matches_ignore_case.group(7)
            quality = matches_ignore_case.group(8)
            comment = (matches_ignore_case.group(9) or '.')[1:]
            extension = matches_ignore_case.group(10)

            recommend = '{}_{}_{}_{}_{}_{}_{}.{}.{}.{}'.format(
                location, year, month, date, track_id, track_number, gender, quality.upper(
                ), comment, extension.lower()
            )
        if not recommend:
            hint += ' has other problem than lower/upper case'

        if song_name_is_correct == 'Yes':
            ws_noproblems.append([
                song_name,
                hint, recommend
            ])
        else:
            ws_problems.append([
                song_name,
                hint, recommend
            ])
            inconsistency_count += 1

        bar.next()
    bar.finish()

    wb.save('Name_pattern_check.xlsx')
    print('Number of inconsistent song names: {} '.format(inconsistency_count))


finally:
    if conn is not None:
        conn.close()
