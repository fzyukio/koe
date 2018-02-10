"""
Import syllables (not elements) from luscinia (after songs have been imported)
"""
import re
from logging import warning

import psycopg2
from django.core.management import BaseCommand
from openpyxl import load_workbook

from root.models import *

name_regex = re.compile('(\w{3})_(\d{4})_(\d{2})_(\d{2})_([\w\d]+)_(\d+)_(\w+)\.(B|EX|VG|G|OK)(\..*)?\.wav')

COLUMN_NAMES = ['Song name', 'Problem', 'Recommend name']

ColumnName = enum(
    POPULATION='Population',
    SONG_NAME='Song Name',
    SYL_START='Syllable Start',
    SYL_END='Syllable End',
    FAMILY='Family',
    SUBFAMILY='Subfamily',
    LABEL='Label'
)

LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

def get_column_letter(col_idx):
    result = []
    while col_idx:
        col_idx, rem = divmod(col_idx - 1, 26)
        result[:0] = LETTERS[rem]
    return ''.join(result)


def get_cell_address(col_idx, row_idx):
    """
    Convert given row and column number to an Excel-style cell name.
    e.g. the first cell is at row 1, col 1, address A1

    :param row_idx: based 1 row index
    :param col_idx: based 1 column index
    :return: address of the cell
    """
    return '{}{}'.format(get_column_letter(col_idx), row_idx)


class Command(BaseCommand):
    def add_arguments(self, parser):

        parser.add_argument(
            '--db',
            action='store',
            dest='db',
            required=True,
            type=str,
            help='Database name',
        )

        parser.add_argument(
            '--port',
            action='store',
            dest='port',
            required=True,
            type=str,
            help='Port',
        )

        parser.add_argument(
            '--host',
            action='store',
            dest='host',
            required=True,
            type=str,
            help='Host',
        )

        parser.add_argument(
            '--xls',
            action='store',
            dest='xls',
            required=True,
            type=str,
            help='Full path to the xls/xlsx/xlsm... file',
        )

        parser.add_argument(
            '--label',
            action='store',
            dest='label',
            required=True,
            type=str,
            help='Path to the label xls',
        )

    def handle(self, db, port, host, xls, label, *args, **options):

        conn = None

        name_wb = load_workbook(filename=xls, read_only=True, data_only=True)
        name_ws = name_wb['Labels']

        label_wb = load_workbook(filename=label, read_only=True, data_only=True)
        label_ws = label_wb['Complete']

        try:
            port = int(port)
            conn = psycopg2.connect("dbname={} user=sa password='sa' host={} port={}".format(db, host, port))
            conn.set_client_encoding('LATIN1')

            cur = conn.cursor()

            cur.execute('select name from songdata')
            results = cur.fetchall()

            old_to_new = {}

            for x in results:
                name = x[0]
                old_to_new[name] = name

            rows = list(name_ws.rows)
            # old_to_new = {}
            # old_cleaned_to_new = {}

            for idx in range(len(rows)):
                if idx == 0:
                    continue
                row = rows[idx]
                original_name = row[0].value
                corrected_name = row[2].value

                old_to_new[original_name] = corrected_name
                # old_cleaned_to_new[original_name.replace(' ', '').replace('$', '')] = corrected_name

                matches = name_regex.match(corrected_name)
                if matches is None:
                    warning('Corrected name at row {} is still incorrect: {}'.format(idx, corrected_name))
                    continue
            name_wb.close()

            old_names = list(old_to_new.keys())
            old_24_to_old = {}
            for name in old_names:
                name_24 = name[:24]
                name_clean = name.replace(' ', '').replace('$', '')
                old_to_new[name_24] = old_to_new[name]
                old_to_new[name_clean] = old_to_new[name]
                old_24_to_old[name_24] = name

            old_names = list(old_to_new.keys())

            col_idx = {cell.value: n for n, cell in enumerate(next(label_ws.rows)) if
                       cell.value in ColumnName.values}
            if len(col_idx) != len(ColumnName.values):
                raise Exception('The excel file does not contain one or more of the mandatory columns: {}'
                                .format(ColumnName.values))

            rows = list(label_ws.rows)
            not_found = {}
            found = {}

            for idx in range(len(rows)):
                if idx == 0:
                    continue
                row = rows[idx]
                song_name = row[col_idx[ColumnName.SONG_NAME]].value
                exists = song_name in old_names

                # if song_name == 'TMI_2015_02_03_MMR052_01_M.VG.Rpu-WM.wav':
                #     print('New name = {}'.format(old_to_new[song_name]))

                if not exists:
                    if song_name not in not_found:
                        not_found[song_name] = []
                    not_found[song_name].append(idx + 1)
                else:
                    if song_name not in found:
                        found[song_name] = []
                    found[song_name].append(idx + 1)

            nlinesfound = 0
            nlinesnotfound = 0
            for name in not_found:
                name_clean = name.replace(' ', '').replace('$', '')
                name_24 = name[:24]

                nlines = len(not_found[name])
                nlinesnotfound += nlines
                # lines = ', '.join(list(map(str,not_found[name])))
                print('Not found: {} in {} lines'.format(name, nlines))
                if name_clean in old_names:
                    print('--->But found as cleaned {}, change to {}'.format(name_clean, old_to_new[name_clean]))
                elif name_24 in old_names:
                    print('--->But found as :24 {}, change to {}'.format(old_24_to_old[name_24], old_to_new[name_24]))

            # for name in found:
            #     nlines = len(found[name])
            #     nlinesfound += nlines
            #     # lines = ', '.join(list(map(str,not_found[name])))
            #     print('Found: {} in {} lines'.format(name, nlines))

            print('-------------------------------')
            print('Total found: {} files, {} lines'.format(len(found.keys()), nlinesfound))
            print('Total not found: {} files, {} lines'.format(len(not_found.keys()), nlinesnotfound))

            name_wb.close()
            label_wb.close()

            label_wb = load_workbook(filename=label, read_only=False, data_only=True)
            label_ws = label_wb['Complete']

            rows = list(label_ws.rows)
            for idx in range(len(rows)):
                if idx == 0:
                    continue
                row = rows[idx]
                cell = row[col_idx[ColumnName.SONG_NAME]]
                name = cell.value
                name_clean = name.replace(' ', '').replace('$', '')
                name_24 = name[:24]

                if name in old_names:
                    new_name = old_to_new[name]
                elif name_clean in old_names:
                    new_name = old_to_new[name_clean]
                elif name_24 in old_names:
                    new_name = old_to_new[name_24]
                else:
                    print('{} not found'.format(name))
                    new_name = name

                cell.value = new_name

            label_wb.save('blah.xlsx')
            label_wb.close()
        finally:
            if conn is not None:
                conn.close()
