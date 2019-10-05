from __future__ import print_function

import array
import io
import os
import pickle
import sys
import traceback
from io import BytesIO
from logging import warning
from pathlib import Path

import numpy as np
import openpyxl
import pydub
from PIL import Image
from django.core.management.base import BaseCommand
from openpyxl.drawing.image import Image as XImage

from koe import wavfile as wf
from root.utils import ensure_parent_folder_exists

PY3 = sys.version_info[0] == 3
if PY3:
    def str_to_bytes(x):
        return str.encode(x, encoding='LATIN9')
else:
    def str_to_bytes(x):
        return x


COLOURS = [[69, 204, 255], [73, 232, 62], [255, 212, 50], [232, 75, 48], [170, 194, 102]]
FF_COLOUR = [0, 0, 0]
AXIS_COLOUR = [127, 127, 127]

COLUMN_NAMES = ['Individual Name', 'Song Id', 'Song Name', 'Syllable Number', 'Syllable Start', 'Syllable End',
                'Syllable Length', 'Element Count', 'Spectrogram', 'Family', 'Label', 'Note', 'Context', 'Min FF', 'Max FF', 'Min Octave',
                'Max Octave', 'Gap Before', 'Gap After', 'Overall instantaneous peak freq', 'Overall peak frequency',
                'Overall Min Freq', 'Overall Max Freq']

LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'


def get_column_letter(col_idx):
    result = []
    while col_idx:
        col_idx, rem = divmod(col_idx - 1, 26)
        result[:0] = LETTERS[rem]
    return ''.join(result)


NOTE_MAP = ['C', 'C#', 'D', 'D#', 'E', 'F', 'F#', 'G', 'G#', 'A', 'A#', 'B']


def hz_to_note(hz):
    midi = 12 * (np.log2(np.atleast_1d(hz)) - np.log2(440.0)) + 69
    note_num = int(np.round(midi))
    note_cents = int(100 * np.around(midi - note_num, 2))

    note = NOTE_MAP[note_num % 12]
    note = '{:s}[{:0d}]'.format(note, int(note_num / 12) - 1)
    note = '{:s} {:+02d}'.format(note, note_cents)

    return note


def get_cell_address(col_idx, row_idx):
    """
    Convert given row and column number to an Excel-style cell name.
    e.g. the first cell is at row 1, col 1, address A1

    :param row_idx: based 1 row index
    :param col_idx: based 1 column index
    :return: address of the cell
    """
    return '{}{}'.format(get_column_letter(col_idx), row_idx)


def get_syllable_end_time(el_rows):
    """
    The reason for this is we can't rely on the last element to find where the syllable actually ends, despite they
    being sorted by starttime. The last element always start later than the element before it but it could be too short
    it finishes before the previous element finishes

    :param el_rows:
    :return:
    """
    end_time = -1
    for row in el_rows:
        el_end_time = row['starttime'] + row['timelength']
        if el_end_time > end_time:
            end_time = el_end_time
    return end_time


def song_id_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    val = song_row['songid']
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def song_individual_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    val = song_row['iname']
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def song_context_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    val = song_row['context']
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def song_name_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    val = song_row['filename']
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_number_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    val = syl_row_idx
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_start_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    val = el_rows[0]['starttime']
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_end_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    val = get_syllable_end_time(el_rows)
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_length_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    syl_starttime = el_rows[0]['starttime']
    syl_endtime = get_syllable_end_time(el_rows)
    val = syl_endtime - syl_starttime
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_element_count_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    val = len(el_rows)
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_spectrogram_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    # wrong_syl_starttime = syl_row['starttime']

    syl_starttime = el_rows[0]['starttime']
    syl_endtime = get_syllable_end_time(el_rows)

    nyquist = song_row['maxfreq']
    fbin = song_row['dy']

    if nyquist == 0:
        nyquist = el_rows[0]['maxf']
    if fbin == 0:
        fbin = el_rows[0]['dy']

    freq_axis_width = 20

    width = int(syl_endtime - syl_starttime) + 1 + freq_axis_width
    height = int(nyquist / fbin)

    img_data_rgb = np.ones((height, width, 3), dtype=np.uint8) * 255
    img_data_bin = np.full((height, width), False, dtype=np.bool)

    for el_idx, el in enumerate(el_rows):
        signal = list(map(int, el['signal'].strip().split(' ')))
        fundfreq = np.array(el['fundfreq'].strip().split(' '), dtype='|S32').astype(np.float) / (nyquist * 2) * height

        # the first 4 numbers of fundfreq are: max, min, ? (no idea) and ? (no idea), so we ignore them
        fundfreq = fundfreq[4:].astype(np.int)
        i = 0
        ff_row_idx = 0
        while i < len(signal):
            num_data = signal[i]
            img_col_idx = signal[i + 1] - syl_starttime + freq_axis_width

            # Draw the mask
            for j in range(2, num_data, 2):
                _signal_segment_end = signal[i + j]
                _signal_segment_start = signal[i + j + 1]
                img_data_rgb[_signal_segment_start:_signal_segment_end, img_col_idx, :] = COLOURS[el_idx % len(COLOURS)]

            # Add the fundamental (red lines)
            if ff_row_idx < len(fundfreq):
                img_row_idx = height - fundfreq[ff_row_idx] - 1

                img_row_idx_padded_low = max(0, img_row_idx - 2)
                img_row_idx_padded_high = img_row_idx + 4 - (img_row_idx - img_row_idx_padded_low)
                img_data_rgb[img_row_idx_padded_low:img_row_idx_padded_high, img_col_idx, :] = FF_COLOUR
                img_data_bin[img_row_idx_padded_low:img_row_idx_padded_high, img_col_idx] = True
            ff_row_idx += 1
            i += (num_data + 1)

    # y_tick_interval = 0 #height // 16
    # x_tick_interval = 0 #10
    # row_idx = 0
    #
    # while row_idx <= height:
    #     if row_idx == 0:
    #         start_row_idx = 0
    #         end_row_idx = 4
    #     elif row_idx == height:
    #         start_row_idx = height - 4
    #         end_row_idx = height
    #     else:
    #         start_row_idx = row_idx - 2
    #         end_row_idx = row_idx + 2
    #
    #     if row_idx in [0, height // 2, height]:
    #         length = freq_axis_width - 10
    #
    #     elif row_idx in [height // 4, height * 3 // 4]:
    #         length = freq_axis_width - 12
    #
    #     elif row_idx in [64, 192, 320, 448]:
    #         length = freq_axis_width - 14
    #
    #     else:
    #         length = freq_axis_width - 16
    #
    #     # print('Idx = {} Row = {}:{}'.format(row_idx, start_row_idx, end_row_idx))
    #     img_data_rgb[start_row_idx:end_row_idx, 0:length] = AXIS_COLOUR
    #     row_idx += y_tick_interval
    #
    # col_idx = freq_axis_width
    # while col_idx <= width:
    #     if col_idx == 0:
    #         start_col_idx = 0
    #         end_col_idx = 1
    #     elif col_idx == height:
    #         start_col_idx = height - 1
    #         end_col_idx = height
    #     else:
    #         start_col_idx = col_idx - 1
    #         end_col_idx = col_idx
    #
    #     img_data_rgb[20:40, start_col_idx:end_col_idx] = AXIS_COLOUR
    #     col_idx += x_tick_interval

    # Draw the actual spectrogram
    # song_name = song_row['filename']
    # song_path = 'user_data/luscinia/luscinia_wav/' + song_name
    #
    # fs = read_wav_info(song_path)
    # spect = extractors['spect'](song_path, fs, syl_starttime, syl_endtime)
    # psd2img(spect, '/tmp/test/img_spect.png')

    # correct_h, correct_w = spect.shape
    # img_h, img_w = img_data_bin.shape

    # from scipy.ndimage.interpolation import zoom
    # img_data_bin = zoom(img_data_bin, (correct_h / img_h, correct_w / img_w))
    # img_data_rgb = zoom(img_data_rgb, (correct_h / img_h, correct_w / img_w, 1))

    # img_bin = Image.fromarray(img_data_bin)
    img_rgb = Image.fromarray(img_data_rgb)

    # np.save('/tmp/luscinia_pkl/' + song_row['filename'][:-4] + 'pkl', np.asarray(img_data_rgb))
    pkl_file_name = '/tmp/luscinia_pkl/' + syl_key_id + ".pkl"

    if not Path.exists(Path(pkl_file_name).parent):
        os.mkdir(Path(pkl_file_name).parent)

    with open(pkl_file_name, 'wb') as f:
        pickle.dump({'binary': img_data_bin}, f, pickle.HIGHEST_PROTOCOL)

    thumbnail_width = int(img_rgb.size[0])
    thumbnail_height = int(img_rgb.size[1] * 0.3)

    img_rgb = img_rgb.resize((thumbnail_width, thumbnail_height))

    output = BytesIO()
    img_rgb.save(output, format='PNG')
    contents = output.getvalue()
    output.close()

    ximg = XImage(io.BytesIO(contents))

    ws.add_image(ximg, get_cell_address(excel_col_idx, excel_row_idx))
    ws.row_dimensions[excel_row_idx].height = thumbnail_height / 1.324137931

    # img_rgb.save('/tmp/test/img_rgb.png', format='PNG')
    # img_bin.save('/tmp/test/img_bin.png', format='PNG')

    return width // 6


def syllable_gapbefore_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    val = el_rows[0]['gapbefore']
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_gapafter_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    val = el_rows[0]['gapafter']
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_minff_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    nyquist = song_row['maxfreq']
    if nyquist == 0:
        nyquist = el_rows[0]['maxf']
    minff = nyquist
    for el in el_rows:
        el_min_ff = float(el['fundfreq'].strip().split(' ')[1])
        minff = min(minff, el_min_ff)

    val = minff
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_maxff_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    maxff = 0
    for el in el_rows:
        el_max_ff = float(el['fundfreq'].strip().split(' ')[0])
        maxff = max(maxff, el_max_ff)

    val = maxff
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_minoctave_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    nyquist = song_row['maxfreq']
    if nyquist == 0:
        nyquist = el_rows[0]['maxf']
    minff = nyquist
    for el in el_rows:
        el_min_ff = float(el['fundfreq'].strip().split(' ')[1])
        minff = min(minff, el_min_ff)

    val = hz_to_note(minff)
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_maxoctave_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    maxff = 0
    for el in el_rows:
        el_max_ff = float(el['fundfreq'].strip().split(' ')[0])
        maxff = max(maxff, el_max_ff)

    val = hz_to_note(maxff)
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_overallpeakfreq1_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    retval = 0
    for el in el_rows:
        el_val = el['overallpeakfreq1']
        retval = max(retval, el_val)
    val = retval
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_overallpeakfreq2_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    retval = 0
    for el in el_rows:
        el_val = el['overallpeakfreq2']
        retval = max(retval, el_val)
    val = retval
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_overallminfreq_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    val = '???'
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def syllable_overallmaxfreq_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    val = '???'
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


def empty_column_handler(song_row, syl_row, el_rows, syl_row_idx, excel_row_idx, excel_col_idx, ws, syl_key_id):
    val = '                                '
    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
    return len(str(val))


COLUMN_HANDLERS = {
    'Song Id': song_id_handler,
    'Individual Name': song_individual_handler,
    'Song Name': song_name_handler,
    'Syllable Number': syllable_number_handler,
    'Syllable Start': syllable_start_handler,
    'Syllable End': syllable_end_handler,
    'Syllable Length': syllable_length_handler,
    'Element Count': syllable_element_count_handler,
    'Spectrogram': syllable_spectrogram_handler,
    'Family': empty_column_handler,
    'Label': empty_column_handler,
    'Note': empty_column_handler,
    'Context': song_context_handler,
    'Gap Before': syllable_gapbefore_handler,
    'Gap After': syllable_gapafter_handler,
    'Min FF': syllable_minff_handler,
    'Max FF': syllable_maxff_handler,
    'Min Octave': syllable_minoctave_handler,
    'Max Octave': syllable_maxoctave_handler,
    'Overall instantaneous peak freq': syllable_overallpeakfreq1_handler,
    'Overall peak frequency': syllable_overallpeakfreq2_handler,
    'Overall Min Freq': syllable_overallminfreq_handler,
    'Overall Max Freq': syllable_overallmaxfreq_handler
}


def export_pcm2wav(song, cur, wav_file_path=None):
    if not os.path.isfile(wav_file_path):
        # print('Importing {}'.format(song_name))
        song_id = song['songid']
        cur.execute('select wav from wavs where songid={};'.format(song_id))

        data = cur.fetchone()
        raw_pcm = str_to_bytes(data[0])

        nchannels = song['stereo']
        bitrate = int(song['ssizeinbits'])
        fs = int(song['samplerate'])

        byte_per_frame = int(bitrate / 8)
        nframes_all_channel = int(len(raw_pcm) / byte_per_frame)
        nframes_per_channel = int(nframes_all_channel / nchannels)
        length = nframes_per_channel
        ensure_parent_folder_exists(wav_file_path)

        if bitrate == 24:
            array1 = np.frombuffer(raw_pcm, dtype=np.ubyte)
            array2 = array1.reshape((nframes_per_channel, nchannels, byte_per_frame)).astype(np.uint8)
            wf.write_24b(wav_file_path, fs, array2)
        else:
            data = array.array('i', raw_pcm)
            sound = pydub.AudioSegment(data=data, sample_width=byte_per_frame, frame_rate=fs, channels=nchannels)
            sound.export(wav_file_path, 'wav')


class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument(
            '--host',
            action='store',
            dest='host',
            default='localhost',
            help='Host address of the H2 - Postgre server',
        )

        parser.add_argument(
            '--port',
            action='store',
            dest='port',
            required=True,
            type=int,
            help='Port of the H2 - Postgre server',
        )

        parser.add_argument(
            '--db',
            action='store',
            dest='db',
            required=True,
            help='Luscinia database name',
        )

    def handle(self, host, port, db, *args, **options):
        import psycopg2.extras

        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Syllable', 0)

        col_max_widths = {}

        ws.append(COLUMN_NAMES)
        for col_name in COLUMN_NAMES:
            col_max_widths[col_name] = len(col_name)

        conn = None

        try:
            conn = psycopg2.connect("dbname={} user=sa password='sa' host={} port={}".format(db, host, port))
            conn.set_client_encoding('LATIN9')

            song_cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            wavs_cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            syl_cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            el_cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            pcm_cur = conn.cursor()

            song_cur.execute('select s.id as songid, s.name as filename, s.maxfreq, s.dy, i.name as iname, '
                             's.call_context as context '
                             'from songdata s '
                             'join individual i on s.individualid=i.id order by i.id ')
            song_row = song_cur.fetchone()
            excel_row_idx = 2

            syl_info = []
            while song_row is not None:
                syl_id = 1

                song_id = song_row['songid']
                print(song_id)
                song_name = song_row['filename']

                wavs_cur.execute('select framesize, stereo, ssizeinbits, samplerate from wavs where songid={}'.format(song_id))
                wavs_row = wavs_cur.fetchone()

                song_row.update(wavs_row)

                wav_file_path = "/tmp/luscinia_wav/" + song_name
                export_pcm2wav(song_row, pcm_cur, wav_file_path)

                syl_cur.execute('select starttime, endtime from syllable where songid={} order by starttime'
                                .format(song_id))
                syl_rows = syl_cur.fetchall()

                for syl_idx, syl_row in enumerate(syl_rows):
                    syl_starttime = syl_row['starttime']
                    syl_endtime = syl_row['endtime']

                    syl_key_id = str(song_id)+'_'+str(syl_id)
                    syl_info_one = [syl_key_id, song_name, syl_starttime, syl_endtime]
                    syl_info.append(syl_info_one)
                    syl_id = syl_id + 1

                    el_cur.execute('select signal, starttime, timelength, fundfreq, gapbefore, gapafter, maxf, dy,'
                                   'overallpeakfreq1, overallpeakfreq2 '
                                   'from element where songid={} and starttime >= {} and (starttime + timelength) <= {}'
                                    .format(song_id, syl_starttime, syl_endtime))
                    el_rows = el_cur.fetchall()

                    if len(el_rows) == 0:
                        warning('Syllable #{} starttime={} endtime={} of song: "{}" doesn\'t enclose any syllable.'
                                ' It will not be put in the spreadsheet. This occurs at row #{}'
                                .format(syl_idx, syl_starttime, syl_endtime, song_name, excel_row_idx))
                        continue

                    for col_idx, col_name in enumerate(COLUMN_NAMES):
                        handler = COLUMN_HANDLERS[col_name]
                        try:
                            width = handler(song_row, syl_row, el_rows, syl_idx + 1, excel_row_idx, col_idx + 1, ws, syl_key_id)
                            if col_max_widths[col_name] < width:
                                col_max_widths[col_name] = width
                        except Exception as e:
                            traceback.print_exc()
                            continue
                    excel_row_idx += 1
                song_row = song_cur.fetchone()

            # Write as a CSV file with headers on first line
            with open('luscinia_info.tsv', 'w') as fp:
                for line in syl_info:
                    fp.write('\t'.join(map(str, line)))
                    fp.write('\n')

        finally:
            if conn is not None:
                conn.close()

            for col_idx, col_name in enumerate(COLUMN_NAMES):
                ws.column_dimensions[get_column_letter(col_idx + 1)].width = col_max_widths[col_name]

            wb.save('export_luscinia_{}_colour.xlsx'.format(db))
