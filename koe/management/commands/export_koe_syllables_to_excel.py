"""
This script will export syllables from the KOE database with added information from Luscinia
"""

from __future__ import print_function

import io

from django.core.management.base import BaseCommand

import numpy as np
import openpyxl
from dotmap import DotMap
from openpyxl.drawing.image import Image as XImage
from PIL import Image

from koe.grid_getters import bulk_get_segment_info
from koe.jsons import tables
from koe.model_utils import get_or_error
from koe.models import Database, Segment
from koe.utils import PAGE_CAPACITY
from root.models import User


COLOURS = [
    [69, 204, 255],
    [73, 232, 62],
    [255, 212, 50],
    [232, 75, 48],
    [170, 194, 102],
]
FF_COLOUR = [0, 0, 0]
AXIS_COLOUR = [127, 127, 127]

LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def get_column_letter(col_idx):
    result = []
    while col_idx:
        col_idx, rem = divmod(col_idx - 1, 26)
        result[:0] = LETTERS[rem]
    return "".join(result)


def get_cell_address(col_idx, row_idx):
    """
    Convert given row and column number to an Excel-style cell name.
    e.g. the first cell is at row 1, col 1, address A1

    :param row_idx: based 1 row index
    :param col_idx: based 1 column index
    :return: address of the cell
    """
    return "{}{}".format(get_column_letter(col_idx), row_idx)


def syllable_spectrogram_handler(image_url, excel_row_idx, excel_col_idx, ws):
    pic = Image.open(image_url)
    img_data_rgb = np.array(pic)

    height, width, _ = np.shape(img_data_rgb)
    freq_axis_width = 20

    y_tick_interval = height // 16
    row_idx = 0

    while row_idx <= height:
        if row_idx == 0:
            start_row_idx = 0
            end_row_idx = 4
        elif row_idx == height:
            start_row_idx = height - 4
            end_row_idx = height
        else:
            start_row_idx = row_idx - 2
            end_row_idx = row_idx + 2

        if row_idx in [0, height // 2, height]:
            length = freq_axis_width - 10

        elif row_idx in [height // 4, height * 3 // 4]:
            length = freq_axis_width - 12

        elif row_idx in [64, 192, 320, 448]:
            length = freq_axis_width - 14

        else:
            length = freq_axis_width - 16

        # print('Idx = {} Row = {}:{}'.format(row_idx, start_row_idx, end_row_idx))
        img_data_rgb[start_row_idx:end_row_idx, 0:length] = AXIS_COLOUR
        row_idx += y_tick_interval

    img = Image.fromarray(img_data_rgb)
    thumbnail_width = int(img.size[0])
    thumbnail_height = int(img.size[1])

    img = img.resize((thumbnail_width, thumbnail_height))

    output = io.BytesIO()
    img.save(output, format="PNG")
    ximg = XImage(output)

    ws.add_image(ximg, get_cell_address(excel_col_idx, excel_row_idx))
    ws.row_dimensions[excel_row_idx].height = thumbnail_height / 1.324137931

    return width / 6


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument(
            "--database-name",
            action="store",
            dest="database_name",
            required=True,
            type=str,
        )
        parser.add_argument(
            "--username",
            action="store",
            dest="username",
            required=True,
            type=str,
        )

    def handle(self, database_name, username, *args, **options):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Syllable", 0)

        database = get_or_error(Database, dict(name__iexact=database_name))
        user = get_or_error(User, dict(username__iexact=username))

        table = tables["segment-info"]
        column_names = []
        slugs = []
        col_max_widths = {}

        for column in table["columns"]:
            column_name = column.get("name", None)
            if column_name is not None:
                if column_name.startswith("_"):
                    continue
                column_names.append(column_name)
                slug = column["slug"]
                slugs.append(slug)
                col_max_widths[slug] = len(column_name)

        slug_to_col_ind = {}
        for ind, slug in enumerate(slugs):
            slug_to_col_ind[slug] = ind + 1

        ws.append(column_names)
        for col_name in column_names:
            col_max_widths[col_name] = len(col_name)

        segments_ids = Segment.objects.filter(audio_file__database=database).values_list("id", flat=True)
        ids, rows = bulk_get_segment_info(segments_ids, DotMap(dict(viewas=user, user=user, database=database.id)))

        excel_row_idx = 1
        for row in rows[:2]:
            excel_row_idx += 1
            for slug, val in row.items():
                excel_col_idx = slug_to_col_ind.get(slug, None)
                if excel_col_idx is None:
                    continue
                if slug == "spectrogram":
                    page = val // PAGE_CAPACITY
                    image_url = "user_data/spect/syllable/{}/{}.png".format(page, val)
                    width = syllable_spectrogram_handler(image_url, excel_row_idx, excel_col_idx, ws)
                else:
                    ws[get_cell_address(excel_col_idx, excel_row_idx)] = val
                    width = len(str(val))
                if col_max_widths[slug] < width:
                    col_max_widths[slug] = width

        for ind, slug in enumerate(col_max_widths):
            ws.column_dimensions[get_column_letter(ind + 1)].width = col_max_widths[slug]

        excel_filename = "export_koe_{}.xlsx".format(database_name)
        wb.save(excel_filename)
